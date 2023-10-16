import numpy as np
import pandas as pd
from glob import glob
import os
import openpyxl as xl
from sklearn.linear_model import LinearRegression
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from utils.EEM import EEmDataset
from utils.water_standard import IndexStandard
import seaborn as sns

class loadData(object):
    def __init__(self,data_path):
        self.data_path = data_path
        self.data = {}
        self.time = {}
        self.standardCurve = {}

        for file in glob(f'{data_path}raw_data/*.xlsx'):
            fileName = os.path.basename(file).split('.')[0]
            print(f'load file: {file}')
            wb = xl.load_workbook(file)

            if fileName in self.time.keys():
                self.time[fileName].append(list(map(int,wb.sheetnames)))
            else:
                self.time[fileName] = list(map(int,wb.sheetnames))

            for time in wb.sheetnames:
                if fileName in ['NO3-N','TN']:
                    data = pd.DataFrame(columns=[1,2,3])
                    origin_data = pd.read_excel(file, sheet_name=time,header=None)
                    for i in range(3):
                        data.iloc[:,i] = origin_data.iloc[:,2*i] - 2*origin_data.iloc[:,2*i+1]
                else:
                    data = pd.read_excel(file, sheet_name=time,header=None)
                if fileName in self.data.keys():
                    self.data[fileName].append(data)
                else:
                    self.data[fileName] = [data]

        for file in glob(f'{data_path}standard_curve/*.xlsx'):
            time = int(os.path.basename(file).split('.')[0])
            wb = xl.load_workbook(file)
            self.standardCurve[time] = {}
            for index in wb.sheetnames:
                print(f'load standard curve: {file}_{index}')
                self.standardCurve[time][index] = self.calculate_standard_curve(file,sheet_name=index)

        self.SCtime = list(self.standardCurve.keys()).sort() if len(list(self.standardCurve.keys()))>1 else list(self.standardCurve.keys())
        self.SCindexes = wb.sheetnames
        self.indexes = list(self.data.keys())
        self.calculate_data()

    def calculate_standard_curve(self,file,sheet_name):
        df = pd.read_excel(file,sheet_name=sheet_name)
        x = np.asarray(df.iloc[:, 1]).reshape(-1, 1)
        y = np.asarray(df.iloc[:, 0]).reshape(-1, 1)
        reg = LinearRegression().fit(x, y)
        return [reg.coef_[0][0], reg.intercept_[0]]

    def calculate_data(self):
        for index in self.SCindexes:
            for i in range(len(self.time[index])):
                time = self.time[index][i]
                Sctime = self.find_proper_time(time)
                self.data[index][i] = self.data[index][i]*self.standardCurve[Sctime][index][0] + self.standardCurve[Sctime][index][1]

    def find_proper_time(self,time):
        pos = 0
        while (self.SCtime[0] < time)&(pos<len(self.SCtime)-1):
            pos+=1

        return self.SCtime[pos]

def check_path(path):
    isExist = os.path.exists(path)
    if not isExist:
        # Create a new directory because it does not exist
        os.makedirs(path)
        print("Successfully created {path}!")

    return None

def time2YMD(time):
    time = str(time)
    return f'{time[:4]}-{time[4:6]}-{time[6:]}'

class LaxDataset(object):
    def __init__(self,data_path, plot = False, plot_indexes = None, plot_time = None, exp_path = None):
        self.data_path = data_path
        self.exp_path = exp_path
        self.dataFrame = self.real_time_data()
        self.raw_data = loadData(self.data_path)
        self.raw_indexes = self.raw_data.indexes

        if self.exp_path:
            check_path(self.exp_path)
            self.EEM_dataset = EEmDataset(self.data_path + 'EEM/', self.exp_path + 'EEM/')
        else:
            self.EEM_dataset = EEmDataset(self.data_path + 'EEM/')

        self.standard_class = dict()
        self.fill_DF()
        if self.exp_path:
            self.dataFrame.to_excel(os.path.join(self.exp_path, 'data.xlsx'), index=False)

        self.indexes = self.dataFrame.columns.tolist()
        self.EEM_indexes = [f'C{i}' for i in range(1, 6)]
        del self.indexes[:3]
        # self.indexes = list(set(self.indexes) - set(self.EEM_indexes))

        if plot:
            self.markers = ['o', '*', 's', 'p', 'h', 'H', '+', 'x', 'D', 'd', '|', '_']
            self.colors = list(mcolors.TABLEAU_COLORS.values())
            if self.exp_path:
                self.plot_path = os.path.join(self.exp_path,'figures/')
                check_path(self.plot_path)
            self.depth_plot(plot_indexes,plot_time)
            self.corr = self.plot_corr()
            self.EEM_time_series_plot(plot_time)
            self.EEM_proportion_plot(plot_time)

    def real_time_data(self):
        data = pd.read_excel(self.data_path+'real_time_data.xlsx')
        return data

    def fill_DF(self):
        for index in self.raw_indexes:
            for i in range(len(self.raw_data.time[index])):
                time = self.raw_data.time[index][i]
                data = self.raw_data.data[index][i]
                for pos in range(len(data)):
                    if len(self.dataFrame[self.dataFrame['Time'] == time]):
                        row = self.dataFrame[(self.dataFrame['Time'] == time)&(self.dataFrame['Pos'] == pos+1)].index[0]
                        self.dataFrame.at[row, index] = np.asarray(data.iloc[pos,:]).mean()
                    else:
                        print(time)
                        print(index)
                        newRow = pd.DataFrame({
                            'Time': time,
                            'Season': self.season(time),
                            'Pos': pos+1,
                            index:np.asarray(data.iloc[pos,:]).mean()
                        })
                        self.dataFrame = pd.concat([self.dataFrame, newRow], ignore_index=True)

        self.dataFrame['TON'] = self.dataFrame['TN'] - self.dataFrame['NH3-N'] - self.dataFrame['NO2-N'] - self.dataFrame['NO3-N']

        for time, val in self.EEM_dataset.intergration.items():
            for pos, values in val.items():
                for i in range(len(values)):

                    if len(self.dataFrame[self.dataFrame['Time'] == time]):
                        row = self.dataFrame[(self.dataFrame['Time'] == time)&(self.dataFrame['Pos'] == pos)].index[0]
                        self.dataFrame.at[row, f'C{i+1}'] = values[i]
                    else:
                        newRow = pd.DataFrame({
                            'Time': time,
                            'Season': self.season(time),
                            'Pos': pos,
                            f'C{i+1}':values[i]
                        })
                        self.dataFrame = pd.concat([self.dataFrame, newRow], ignore_index=True)

        FL_indexes = ['FI', 'BIX', 'HIX']
        for time, val in self.EEM_dataset.FL_indexes.items():
            for pos, values in val.items():
                for i, val in enumerate(values):

                    if len(self.dataFrame[self.dataFrame['Time'] == time]):
                        row = self.dataFrame[(self.dataFrame['Time'] == time)&(self.dataFrame['Pos'] == pos)].index.item()
                        self.dataFrame.at[row, FL_indexes[i]] = val
                    else:
                        newRow = pd.DataFrame({
                            'Time': time,
                            'Season': self.season(time),
                            'Pos': pos,
                            FL_indexes[i]:val
                        })
                        self.dataFrame = pd.concat([self.dataFrame, newRow], ignore_index=True)

        return None

    def season(self,time):
        if time<20230901:
            return 'summer'
        elif time <20231201:
            return 'autumn'
        elif time <20240301:
            return 'winter'
        else:
            return 'spring'

    def disinfection(self,times = None):
        if times == None:
            times = []
            all_time = [t for t in self.dataFrame['Time'].tolist()]
            all_time = list(dict.fromkeys(all_time))
            for t in all_time:
                if (~np.isnan(self.dataFrame[(self.dataFrame['Time'] == t) & (self.dataFrame['Pos'] == 1)]['NH3-N'].tolist()[0])) & (~np.isnan(self.dataFrame[(self.dataFrame['Time'] == t) & (self.dataFrame['Pos'] == 1)]['DOC'].tolist()[0])):
                    times.append(int(t))

        c = 1000
        print(f'The concentration of Cl is {c} mg/L.')

        for time in times:
            data = self.dataFrame[self.dataFrame['Time']==time]
            for pos in range(1,7):
                DOC = data[data['Pos']==pos]['DOC'].tolist()[0]
                NH3 = data[data['Pos']==pos]['NH3-N'].tolist()[0]
                Cl = 7.6*NH3 + 3*DOC + 10
                V = 50*Cl/c
                print(f'Sample time: {time}, Position: {pos}, volumn: 50 mL, chlorine volumn: {V} mL')

        return None

    def depth_plot(self,indexes = None, times = None):
        if indexes == None:
            indexes = self.indexes
        if times == None:
            times = self.dataFrame['Time'].tolist()
            times = list(dict.fromkeys(times))

        Standard_Generator = IndexStandard(f'{self.data_path}water_standard/water_standard.xlsx')
        dataFrame = self.dataFrame[self.dataFrame['Time'].isin(times)]
        for index in indexes:
            surface_range, surface_indexes, drinking_range, surface_cls = Standard_Generator.get_class(index,dataFrame[index].tolist())
            self.standard_class[index] = surface_cls
            i = 0
            fig, (ax1, ax2) = plt.subplots(1, 2)
            for ind, ax in enumerate([ax1, ax2]):
                ax.tick_params(top=True, labeltop=True, bottom=False, labelbottom=False)
                ax.xaxis.set_label_position('top')
                ax.set_xlabel(f'{index} (unit)')
                ax.set_title(f'Point {ind+1}',y=-0.1)
                self.plot_standard(surface_range, surface_indexes,drinking_range, ax)
            ax1.invert_yaxis()
            ax1.set_ylabel("Depth (cm)")
            ax2.sharey(ax1)
            ax2.yaxis.set_visible(False)  # same for y axis.

            for time in times:
                plot_data_frame = dataFrame[dataFrame['Time']==time]
                i += 1
                if all(plot_data_frame[index].tolist()):
                    self.one_depth_plot(plot_data_frame,index,time, ax1,ax2,i)
                else:
                    continue

            fig.legend(*ax2.get_legend_handles_labels(),
                       loc='upper center',bbox_to_anchor=(1.08, 0.88), borderaxespad=0)
            if self.plot_path:
                fig.savefig(f'{self.plot_path}{index}.png', bbox_inches='tight')
            fig.show()

        return None

    def plot_standard(self, surface_range, surface_indexes, drinking_range, ax):
        if surface_range != None:
            for i, s_r in enumerate(surface_range):
                if i ==0:
                    ax.axvline(x=s_r, color = 'black', ls='--', dashes = (5, 10), label = 'Surface Water Standard')
                else:
                    ax.axvline(x=s_r, color='black', ls='--', dashes = (5, 10))

                ax.text(s_r * 1, 160, surface_indexes[i])

        if drinking_range != None:
            for d_r in drinking_range:
                ax.axvline(x=d_r, color='gray', ls=':', dashes = (5, 10), label='Drinking Water Limit')

    def one_depth_plot(self, dataFrame, index,time, ax1,ax2, i):
        x1 = dataFrame[index].tolist()[:3]
        x2 = dataFrame[index].tolist()[3:]

        depth = {
            1: 30,
            2: 80,
            3: 130,
            4: 30,
            5: 100,
            6: 150
        }

        y1 = [depth[x] for x in dataFrame['Pos'].tolist()[:3]]
        y2 = [depth[x] for x in dataFrame['Pos'].tolist()[3:]]
        ax1.plot(x1, y1, marker=self.markers[i],label = time2YMD(time))
        ax2.plot(x2, y2, marker=self.markers[i],label = time2YMD(time))

        return None

    def EEM_time_series_plot(self, times=None):
        indexes = self.EEM_indexes
        if times == None:
            times = self.EEM_dataset.times

        print(f'times: {times}')

        locations = ['Surface','Middle','Bottom']
        surface_1 = {}
        middle_1 = {}
        bottom_1 = {}
        surface_2 = {}
        middle_2 = {}
        bottom_2 = {}
        for index in indexes:
            surface_1[index] = []
            middle_1[index] = []
            bottom_1[index] = []
            surface_2[index] = []
            middle_2[index] = []
            bottom_2[index] = []

        for i, index in enumerate(indexes):
            for time in times:
                surface_1[index].append(self.EEM_dataset.intergration[time][1][i])
                middle_1[index].append(self.EEM_dataset.intergration[time][2][i])
                bottom_1[index].append(self.EEM_dataset.intergration[time][3][i])
                surface_2[index].append(self.EEM_dataset.intergration[time][4][i])
                middle_2[index].append(self.EEM_dataset.intergration[time][5][i])
                bottom_2[index].append(self.EEM_dataset.intergration[time][6][i])

        fig, ax = plt.subplots(len(indexes), 2, figsize=(20,20))

        x = np.arange(len(times))  # the label locations
        width = 0.25  # the width of the bars
        colors = ["#44a5c2", "#ffae49", "#64CCC5", "#D8B4F8", "#FFB7B7"]

        for i in range(len(indexes)):
            for j in range(2):
                if j == 0:
                    data = [surface_1,middle_1,bottom_1]
                else:
                    data = [surface_2,middle_2,bottom_2]

                for ind, d in enumerate(data):
                    offset = width*ind
                    ax[i,j].bar(x+offset,d[f'C{i+1}'],width,color = colors[ind],edgecolor = 'black',label = locations[ind])

                ax[i,j].set_ylabel('Intergration')
                ax[i,j].set_title(f'Point {j+1}: C{i+1}')
                ax[i,j].set_xticks(x + width, times)
        ax[0,0].legend(loc='upper right', ncols=len(locations))
        fig.savefig(self.plot_path+'EEM_time_series.png')
        plt.show()

        return None

    def EEM_proportion_plot(self, times=None):
        indexes = self.EEM_indexes
        if times == None:
            times = self.EEM_dataset.times

        dataFrame = self.dataFrame[self.dataFrame['Time'].isin(times)]

        groups = ['Surface','Middle','Bottom']
        colors = ["#44a5c2", "#ffae49", "#64CCC5", "#D8B4F8", "#FFB7B7"]

        fig, ax = plt.subplots(len(times),2, figsize=(10,10))

        plt.rc('axes', titlesize=20)  # Controls Axes Title
        plt.rc('axes', labelsize=20)  # Controls Axes Labels
        plt.rc('xtick', labelsize=20)  # Controls x Tick Labels
        plt.rc('ytick', labelsize=20)  # Controls y Tick Labels
        plt.rc('legend', fontsize=10)  # Controls Legend Font
        plt.rc('figure', titlesize=20)  # Controls Figure Title

        for i_time, time in enumerate(times):
            for j in range(2):
                if j == 0:
                    pos = [1,2,3]
                else:
                    pos = [4,5,6]

                values = np.zeros((len(indexes),len(pos)))

                for ind in range(len(indexes)):
                    for k, p in enumerate(pos):
                        values[ind,k] = self.EEM_dataset.proportion[time][p][ind]

                # Stacked bar chart with loop
                for i in range(values.shape[0]):
                    ax[i_time,j].bar(groups, values[i], edgecolor='black', color=colors[i], bottom=np.sum(values[:i], axis=0),
                           label=indexes[i])

                # Labels
                for bar in ax[i_time,j].patches:
                    ax[i_time,j].text(bar.get_x() + bar.get_width() / 2,
                            bar.get_height() / 2 + bar.get_y(),
                            f'{round(bar.get_height()*100, 1)}%', ha='center',
                            color='w', weight='bold', size=10)


            ax[i_time,0].set_ylabel(f'{str(time)[:4]}-{str(time)[4:6]}-{str(time)[6:]}')
        fig.legend(*ax[0,1].get_legend_handles_labels(),
                   loc='upper center', bbox_to_anchor=(0.96, 0.88), borderaxespad=0)
        fig.savefig(self.plot_path + 'EEM_proportion.png', bbox_inches='tight')
        fig.show()

    def plot_corr(self):
        dataFrame = self.dataFrame.iloc[:, 3:]
        corr = dataFrame.corr()
        fig = plt.figure(figsize=(14, 8))
        sns.set_theme(style="white")
        sns.heatmap(corr, annot=True, cmap="RdBu_r", fmt='.1g')
        fig.savefig(f'{self.plot_path}correlation.png')

        return corr
